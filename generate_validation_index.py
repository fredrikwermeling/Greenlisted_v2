#!/usr/bin/env python3
"""
Generates species-specific sgRNA validation index files from all built-in CRISPR libraries.
  - libraries/sgRNA_validation_index_human.txt
  - libraries/sgRNA_validation_index_mouse.txt

This index is used by the Validate sgRNA feature for reverse lookup (sgRNA → gene).
Sequences sourced from Addgene (addgene.org) and Broad Institute GPP
(portals.broadinstitute.org/gpp/public/pool/index).

Usage: python3 generate_validation_index.py
"""

import json
import os
import re
import sys
from itertools import combinations

try:
    import openpyxl
except ImportError:
    print("openpyxl is required: pip3 install openpyxl", file=sys.stderr)
    sys.exit(1)

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
SETTINGS_FILE = os.path.join(SCRIPT_DIR, "settingsLibraries.json")
OUTPUT_FILE_HUMAN = os.path.join(SCRIPT_DIR, "libraries", "sgRNA_validation_index_human.txt")
OUTPUT_FILE_MOUSE = os.path.join(SCRIPT_DIR, "libraries", "sgRNA_validation_index_mouse.txt")
OUTPUT_UPSET_JSON = os.path.join(SCRIPT_DIR, "libraries", "upset_data.json")

# Libraries classified by species (based on synonymName in settingsLibraries.json)
HUMAN_LIBRARIES = {
    "Brunello (human)",
    "GeCKO v2 (human) A+B",
    "Gattinara (human)",
    "Jacquere (human)",
    "VBC (human)",
}

MOUSE_LIBRARIES = {
    "Brie (mouse)",
    "GeCKO v2 (mouse) A+B",
    "Gouda (mouse)",
    "Julianna (mouse)",
    "VBC (mouse)",
}

# Validation-only libraries (XLSX files from Addgene, not used for screen design)
VALIDATION_ONLY_XLSX = [
    {
        "name": "Yusa Human v1",
        "species": "human",
        "file": "libraries/yusa_human_v1_raw.xlsx",
        "gene_col": "Gene",
        "seq_col": "Guide_sequence",
    },
    {
        "name": "Yusa Mouse v2",
        "species": "mouse",
        "file": "libraries/yusa_mouse_v2_raw.xlsx",
        "gene_col": "gene",
        "seq_col": "guide_sequence",
    },
    {
        "name": "TKOv3 (human)",
        "species": "human",
        "file": "libraries/tkov3_raw.xlsx",
        "gene_col": "GENE",
        "seq_col": "SEQUENCE",
    },
    {
        "name": "mTKO (mouse)",
        "species": "mouse",
        "file": "libraries/mtko_raw.xlsx",
        "gene_col": "GENE",
        "seq_col": "SEQUENCE",
    },
    {
        "name": "MinLibCas9 (human)",
        "species": "human",
        "file": "libraries/minlibcas9_raw.xlsx",
        "gene_col": "Approved_Symbol",
        "seq_col": "WGE_Sequence",
        "trim_pam": 3,  # sequence includes 3bp PAM at end
    },
]

# Map of known score column header names
SCORE_HEADERS = {
    "Rule Set 2 score",
    "On-Target Efficacy Score",
    "Aggregate CFD Score",
    "VBC score",
}


def detect_score_columns(headers):
    """Return list of (column_index, header_name) for recognized score columns."""
    result = []
    for i, h in enumerate(headers):
        if h.strip() in SCORE_HEADERS:
            result.append((i, h.strip()))
    return result


def detect_gene_id_column(headers, symbol_col_idx, rna_col_idx):
    """Try to find a Gene ID column from the headers."""
    id_keywords = ["gene id", "gene_id", "target gene id", "annotated gene id"]
    for i, h in enumerate(headers):
        if h.strip().lower() in id_keywords and i != symbol_col_idx and i != rna_col_idx:
            return i
    # Fallback: for Brie/Brunello, column 0 is "Target Gene ID"
    for i, h in enumerate(headers):
        if "gene id" in h.strip().lower() and i != symbol_col_idx and i != rna_col_idx:
            return i
    return None


def process_library(lib_config):
    """Process a single library file and return list of index rows."""
    name = lib_config["name"]
    filepath = os.path.join(SCRIPT_DIR, lib_config["fileName"])
    # settingsLibraries.json uses 1-based column indices
    symbol_col = lib_config["symbolColumn"] - 1
    rna_col = lib_config["RNAColumn"] - 1

    if not os.path.exists(filepath):
        print(f"  WARNING: File not found: {filepath}", file=sys.stderr)
        return []

    rows = []
    with open(filepath, "r", encoding="utf-8") as f:
        header_line = f.readline().rstrip("\n\r")
        headers = header_line.split("\t")
        score_cols = detect_score_columns(headers)
        gene_id_col = detect_gene_id_column(headers, symbol_col, rna_col)

        line_num = 1
        for line in f:
            line_num += 1
            line = line.rstrip("\n\r")
            if not line:
                continue
            cols = line.split("\t")
            if len(cols) <= max(symbol_col, rna_col):
                continue

            sgrna = cols[rna_col].strip()
            symbol = cols[symbol_col].strip()
            gene_id = cols[gene_id_col].strip() if gene_id_col is not None and gene_id_col < len(cols) else ""

            # Build scores string
            score_parts = []
            for sc_idx, sc_name in score_cols:
                if sc_idx < len(cols):
                    val = cols[sc_idx].strip()
                    if val:
                        score_parts.append(f"{sc_name}: {val}")
            scores = "; ".join(score_parts)

            rows.append(f"{sgrna}\t{name}\t{symbol}\t{gene_id}\t{scores}")

    return rows


def process_xlsx_library(config):
    """Process a validation-only XLSX library file and return list of index rows."""
    name = config["name"]
    filepath = os.path.join(SCRIPT_DIR, config["file"])

    if not os.path.exists(filepath):
        print(f"  WARNING: File not found: {filepath}", file=sys.stderr)
        return []

    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb[wb.sheetnames[0]]

    # Find column indices from header row
    headers = None
    gene_idx = None
    seq_idx = None
    rows = []

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(c).strip() if c else "" for c in row]
            for j, h in enumerate(headers):
                if h == config["gene_col"]:
                    gene_idx = j
                if h == config["seq_col"]:
                    seq_idx = j
            if gene_idx is None or seq_idx is None:
                print(f"  WARNING: Could not find columns '{config['gene_col']}'/'{config['seq_col']}' in {headers}", file=sys.stderr)
                return []
            continue

        gene = str(row[gene_idx]).strip() if row[gene_idx] else ""
        seq = str(row[seq_idx]).strip() if row[seq_idx] else ""
        if gene and seq:
            trim_pam = config.get("trim_pam", 0)
            if trim_pam:
                seq = seq[:-trim_pam]
            rows.append(f"{seq}\t{name}\t{gene}\t\t")

    wb.close()
    return rows


def write_index(rows, output_path, label):
    """Write index rows to a file."""
    with open(output_path, "w", encoding="utf-8") as f:
        f.write("# Sequences sourced from Addgene (addgene.org) and Broad Institute GPP (portals.broadinstitute.org/gpp/public/pool/index)\n")
        f.write("sgRNA Sequence\tLibrary\tGene Symbol\tGene ID\tScores\n")
        for row in rows:
            f.write(row + "\n")
    size_mb = os.path.getsize(output_path) / (1024 * 1024)
    print(f"  {label}: {len(rows)} entries, {size_mb:.1f} MB → {output_path}")


def _clean_library_name(name):
    """Remove species suffix like '(human)' or '(mouse)' from a library name."""
    return re.sub(r"\s*\((?:human|mouse)\)\s*", " ", name).strip()


def _collect_from_tsv(lib_config):
    """Return (set_of_sgRNAs, dict_of_gene_to_sgrna_set) from a TSV library file."""
    filepath = os.path.join(SCRIPT_DIR, lib_config["fileName"])
    rna_col = lib_config["RNAColumn"] - 1
    symbol_col = lib_config["symbolColumn"] - 1
    seqs = set()
    gene_map = {}
    if not os.path.exists(filepath):
        return seqs, gene_map
    with open(filepath, "r", encoding="utf-8") as f:
        f.readline()  # skip header
        for line in f:
            line = line.rstrip("\n\r")
            if not line:
                continue
            cols = line.split("\t")
            if len(cols) > max(rna_col, symbol_col):
                seq = cols[rna_col].strip()
                gene = cols[symbol_col].strip()
                if seq:
                    seqs.add(seq)
                    if gene:
                        if gene not in gene_map:
                            gene_map[gene] = set()
                        gene_map[gene].add(seq)
    return seqs, gene_map


def _collect_from_xlsx(config):
    """Return (set_of_sgRNAs, dict_of_gene_to_sgrna_set) from an XLSX library."""
    filepath = os.path.join(SCRIPT_DIR, config["file"])
    if not os.path.exists(filepath):
        return set(), {}
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb[wb.sheetnames[0]]
    seqs = set()
    gene_map = {}
    seq_idx = None
    gene_idx = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(c).strip() if c else "" for c in row]
            for j, h in enumerate(headers):
                if h == config["seq_col"]:
                    seq_idx = j
                if h == config["gene_col"]:
                    gene_idx = j
            if seq_idx is None or gene_idx is None:
                wb.close()
                return seqs, gene_map
            continue
        seq = str(row[seq_idx]).strip() if row[seq_idx] else ""
        gene = str(row[gene_idx]).strip() if row[gene_idx] else ""
        if seq:
            trim_pam = config.get("trim_pam", 0)
            if trim_pam:
                seq = seq[:-trim_pam]
            seqs.add(seq)
            if gene:
                if gene not in gene_map:
                    gene_map[gene] = set()
                gene_map[gene].add(seq)
    wb.close()
    return seqs, gene_map


def _compute_gene_stats(gene_map):
    """Compute gene statistics from a gene→sgRNA_set map."""
    if not gene_map:
        return {"genes": 0, "sgrnas_per_gene": 0}
    counts = [len(sgrnas) for sgrnas in gene_map.values()]
    n_genes = len(gene_map)
    median_idx = n_genes // 2
    sorted_counts = sorted(counts)
    if n_genes % 2 == 1:
        median_val = sorted_counts[median_idx]
    else:
        median_val = (sorted_counts[median_idx - 1] + sorted_counts[median_idx]) / 2
    # Use int if whole number, otherwise one decimal
    median_val = int(median_val) if median_val == int(median_val) else round(median_val, 1)
    return {"genes": n_genes, "sgrnas_per_gene": median_val}


# Target type descriptions per library (cleaned name → description)
LIBRARY_TARGETS = {
    "Brunello": "Protein-coding genes",
    "Brie": "Protein-coding genes",
    "GeCKO v2 A+B": "Protein-coding genes + miRNAs",
    "Gattinara": "Protein-coding genes",
    "Gouda": "Protein-coding genes",
    "Jacquere": "Protein-coding genes",
    "Julianna": "Protein-coding genes",
    "VBC": "Protein-coding genes",
    "Yusa Human v1": "Protein-coding genes",
    "Yusa Mouse v2": "Protein-coding genes",
    "TKOv3": "Protein-coding genes",
    "mTKO": "Protein-coding genes",
    "MinLibCas9": "Protein-coding genes",
}


def generate_upset_data(libraries):
    """Generate UpSet plot data (exclusive intersection counts) for both species."""
    # Collect sgRNA sets and gene maps per library, grouped by species
    species_libs = {"human": [], "mouse": []}

    for lib in libraries:
        name = lib["name"]
        if name in HUMAN_LIBRARIES:
            species = "human"
        elif name in MOUSE_LIBRARIES:
            species = "mouse"
        else:
            continue
        seqs, gene_map = _collect_from_tsv(lib)
        species_libs[species].append((_clean_library_name(name), seqs, gene_map))

    for config in VALIDATION_ONLY_XLSX:
        species = config["species"]
        seqs, gene_map = _collect_from_xlsx(config)
        species_libs[species].append((_clean_library_name(config["name"]), seqs, gene_map))

    result = {}
    for species, lib_list in species_libs.items():
        n = len(lib_list)
        sets_info = []
        for name, seqs, gene_map in lib_list:
            stats = _compute_gene_stats(gene_map)
            sets_info.append({
                "name": name,
                "size": len(seqs),
                "genes": stats["genes"],
                "sgrnas_per_gene": stats["sgrnas_per_gene"],
                "targets": LIBRARY_TARGETS.get(name, "Protein-coding genes"),
            })

        # Build sgRNA → frozenset of library indices
        sgrna_to_libs = {}
        for idx, (name, seqs, gene_map) in enumerate(lib_list):
            for seq in seqs:
                if seq not in sgrna_to_libs:
                    sgrna_to_libs[seq] = set()
                sgrna_to_libs[seq].add(idx)

        # Count each exclusive combination
        combo_counts = {}
        for seq, lib_indices in sgrna_to_libs.items():
            key = frozenset(lib_indices)
            combo_counts[key] = combo_counts.get(key, 0) + 1

        # Convert to sorted list
        intersections = []
        for key, count in combo_counts.items():
            intersections.append({"sets": sorted(key), "size": count})
        intersections.sort(key=lambda x: x["size"], reverse=True)

        result[species] = {"sets": sets_info, "intersections": intersections}
        print(f"  UpSet {species}: {len(sets_info)} libraries, {len(intersections)} exclusive intersections")

    with open(OUTPUT_UPSET_JSON, "w", encoding="utf-8") as f:
        json.dump(result, f, separators=(",", ":"))
    size_kb = os.path.getsize(OUTPUT_UPSET_JSON) / 1024
    print(f"  UpSet data: {size_kb:.1f} KB → {OUTPUT_UPSET_JSON}")


def main():
    with open(SETTINGS_FILE, "r", encoding="utf-8") as f:
        libraries = json.load(f)

    print(f"Processing {len(libraries)} libraries...")

    human_rows = []
    mouse_rows = []
    for lib in libraries:
        name = lib["name"]
        print(f"  {name}...")
        lib_rows = process_library(lib)
        print(f"    → {len(lib_rows)} sgRNAs")

        if name in HUMAN_LIBRARIES:
            human_rows.extend(lib_rows)
        elif name in MOUSE_LIBRARIES:
            mouse_rows.extend(lib_rows)
        else:
            print(f"    WARNING: '{name}' not classified as human or mouse, skipping", file=sys.stderr)

    # Process validation-only XLSX libraries
    print(f"\nProcessing {len(VALIDATION_ONLY_XLSX)} validation-only libraries...")
    for config in VALIDATION_ONLY_XLSX:
        name = config["name"]
        print(f"  {name}...")
        lib_rows = process_xlsx_library(config)
        print(f"    → {len(lib_rows)} sgRNAs")
        if config["species"] == "human":
            human_rows.extend(lib_rows)
        else:
            mouse_rows.extend(lib_rows)

    # Write species-specific outputs
    print()
    write_index(human_rows, OUTPUT_FILE_HUMAN, "Human")
    write_index(mouse_rows, OUTPUT_FILE_MOUSE, "Mouse")

    # Generate UpSet plot data
    print("\nGenerating UpSet plot data...")
    generate_upset_data(libraries)

    # Remove old combined file if it exists
    old_combined = os.path.join(SCRIPT_DIR, "libraries", "sgRNA_validation_index.txt")
    if os.path.exists(old_combined):
        os.remove(old_combined)
        print(f"\nRemoved old combined index: {old_combined}")

    print(f"\nDone. Total: {len(human_rows)} human + {len(mouse_rows)} mouse = {len(human_rows) + len(mouse_rows)} entries")


if __name__ == "__main__":
    main()
